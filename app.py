"""
DDS Allada  —  Сравнение расходов по периодам.

Flask-приложение: выбор периода (месяц / квартал / год),
таблица сравнения год-к-году в браузере + скачивание XLSX.

Выручка берётся из строки 18 листа ДДС
  «Поступления от клиентов (продажа товаров/услуг)»
Кредитные операции выносятся в отдельный блок внизу.

python app.py  →  http://localhost:5000
"""
from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from pathlib import Path
from werkzeug.utils import secure_filename
from datetime import datetime
import calendar, io, warnings, re
warnings.filterwarnings('ignore')

# Google Sheets (опционально — устанавливается только при наличии credentials.json)
try:
    import gspread
    from google.oauth2.service_account import Credentials as _GSCredentials
    _GSPREAD_AVAILABLE = True
except ImportError:
    _GSPREAD_AVAILABLE = False

app = Flask(__name__)


@app.after_request
def add_cors_headers(resp):
    """
    Нужен для сценария, когда UI открыт как file://.../index.html
    и обращается к API на http://127.0.0.1:5000.
    """
    if request.path.startswith('/api/'):
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return resp

# ── Config ─────────────────────────────────────────────────────────────────────
XLSX_IN = None
TARGET_FILE_NAME = "system_dds_allada.xlsx"
ALLOWED_EXTENSIONS = {".xlsx"}

# ── Google Sheets config ───────────────────────────────────────────────────────
GSHEET_ID        = '1TcSFMRqq8zguBJwsqz0-BiDLqTjjpx3JYWdUAMi9izs'
CREDENTIALS_FILE = Path(__file__).resolve().parent / 'credentials.json'
USE_GSHEETS      = False   # включается автоматически при наличии credentials.json
_GS_CLIENT       = None

BANK_WALLETS = {'Альфа 796', 'Сбер 595', 'Счет 606', 'Счет 11', 'Счет 12'}
EXCLUDE_ARTICLES = {
    'Доход — Перевод между счетами',  'Расход — Перевод между счетами',
    'Доход — перевод между счетами',  'Расход — перевод между счетами',
}
RENAME_ARTICLES = {'Реклама/Маркетинг Звягин': 'Привлечение по ДМС Звягин'}

# Статьи, относящиеся к кредитам/займам (не расходы и не доходы)
CREDIT_ARTICLES_SET = {
    'Возврат кредитов и займов',
    'Получение кредитов и займов',
    'Полученные проценты за выданные кредиты и займы',
    'Проценты по кредиту',
}

# Статьи, которые всегда идут в конец административных расходов
_ADMIN_TAIL = ('дивиденды', 'енп')

def _is_admin_tail(label: str) -> bool:
    ll = label.lower()
    return any(t in ll for t in _ADMIN_TAIL)


# Ключевые слова для определения производственных расходов
_PRODUCTION_KW = (
    'стоматолог', 'имплант', 'биоматери', 'брекет', 'элайнер',
    'лаборатор', 'сэс', 'роспотребнадзор', 'спецодежд',
    'профосмотр', 'профсмотр', 'закуп', 'фонд оплаты',
    'налоги и взносы', 'производственн', 'прочее оборудование',
    'ремонт стоматол', 'доставка материал',
    'реклам', 'маркетинг', 'привлечение',
)

def get_expense_category(article_name: str) -> str:
    """Производственные или Административные расходы."""
    a = article_name.lower()
    for kw in _PRODUCTION_KW:
        if kw in a:
            return 'Производственные'
    return 'Административные'

def _get_grp_category(arts: list) -> str:
    for a in arts:
        if get_expense_category(a) == 'Производственные':
            return 'Производственные'
    return 'Административные'


def is_revenue_article(name: str) -> bool:
    """True для статей выручки (поступления, возвраты за услуги).
    Эти статьи не включаются в расходы — они уже отражены в строке 18 ДДС.
    """
    n = name.lower()
    if 'поступлен' in n:
        return True
    # Возврат за услуги (но не возврат кредитов/займов)
    if 'возврат' in n and ('услуг' in n or 'стоматол' in n):
        return True
    return False


def _rev_val(article: str, amount: float) -> float:
    """Знако-корректированное значение для разбивки выручки.
    Поступления → положительное; Возвраты (уменьшение выручки) → отрицательное.
    """
    if 'возврат' in article.lower():
        return -abs(amount)
    return abs(amount)

JOURNAL_SHEETS = [
    ('Журнал Снабжение',   0, 2, 6, 1),
    ('Журнал МобиДик',     0, 2, 6, 1),
    ('Журнал Леонова',     0, 2, 6, 1),
    ('Журнал Мира 100',    0, 2, 6, 1),
    ('Журнал Таштемирова', 0, 2, 6, 1),
    ('Журнал Ирина Влад.', 0, 2, 6, 1),
]
BANK_SHEETS = [('1С Альфа', 0, 12, 13, 0), ('1С Сбер', 0, 12, 13, 0)]

MONTH_RU = {1:'Январь',2:'Февраль',3:'Март',4:'Апрель',5:'Май',6:'Июнь',
            7:'Июль',8:'Август',9:'Сентябрь',10:'Октябрь',11:'Ноябрь',12:'Декабрь'}

# ── Маппинг колонок листа ДДС ─────────────────────────────────────────────────
# ДДС row 18 (0-indexed 17) = "Поступления от клиентов (продажа товаров/услуг):"
# Pandas col offsets (0-indexed) per year:
#   2024: month M → col M        (Jan=1, Dec=12)
#   2025: month M → col 14 + M   (Jan=15, Dec=26)
#   2026: month M → col 31 + M   (Jan=32, Dec=43)
DDS_REVENUE_ROW = 17    # pandas 0-indexed (Excel row 18)
DDS_YEAR_OFFSETS = {2024: 0, 2025: 14, 2026: 31}

# ── Data loading (cached) ─────────────────────────────────────────────────────
_TX = None
_DDS = None


def reset_cache():
    global _TX, _DDS, _GS_CLIENT
    _TX = None
    _DDS = None
    _GS_CLIENT = None


# ── Google Sheets helpers ──────────────────────────────────────────────────────
def _get_gs_client():
    """Создаёт (и кэширует) клиент gspread через Service Account.
    Credentials берутся из credentials.json (локально) или
    из переменной окружения GOOGLE_CREDENTIALS_JSON (продакшн/Render).
    """
    global _GS_CLIENT
    if _GS_CLIENT is None:
        if not _GSPREAD_AVAILABLE:
            raise RuntimeError(
                'Пакеты gspread/google-auth не установлены. '
                'Выполните: pip install gspread google-auth'
            )
        import json, os, tempfile
        scopes = ['https://www.googleapis.com/auth/spreadsheets.readonly']

        if CREDENTIALS_FILE.exists():
            # Локальный запуск — читаем файл напрямую
            creds = _GSCredentials.from_service_account_file(str(CREDENTIALS_FILE), scopes=scopes)
        else:
            # Облачный запуск — берём JSON из переменной окружения
            creds_json = os.environ.get('GOOGLE_CREDENTIALS_JSON')
            if not creds_json:
                raise RuntimeError(
                    'credentials.json не найден и переменная окружения '
                    'GOOGLE_CREDENTIALS_JSON не задана.'
                )
            creds_info = json.loads(creds_json)
            creds = _GSCredentials.from_service_account_info(creds_info, scopes=scopes)

        _GS_CLIENT = gspread.authorize(creds)
    return _GS_CLIENT


def _clean_number(v):
    """Конвертирует строку из Google Sheets в число (учитывает пробел-разделитель тысяч)."""
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace('\xa0', '').replace('\u2009', '').replace(' ', '')
    s = s.replace(',', '.')
    try:
        return float(s)
    except (ValueError, TypeError):
        return float('nan')


def _load_sheet_gs(gc, name, dc, ac, rc, hr, wc=None):
    """Читает лист из Google Sheets и возвращает DataFrame транзакций."""
    try:
        sh = gc.open_by_key(GSHEET_ID)
        ws = sh.worksheet(name)
        data = ws.get_all_values()
    except Exception as e:
        print(f"  Лист '{name}' пропущен: {e}")
        return pd.DataFrame()

    if not data or len(data) <= hr:
        return pd.DataFrame()

    headers = data[hr]
    rows    = data[hr + 1:]
    df = pd.DataFrame(rows, columns=headers)

    cols = df.columns.tolist()
    if max(dc, ac, rc) >= len(cols):
        return pd.DataFrame()

    df = df.rename(columns={cols[dc]: 'date', cols[ac]: 'amount', cols[rc]: 'article'})
    keep = ['date', 'amount', 'article']
    if wc is not None and wc < len(cols):
        df = df.rename(columns={cols[wc]: 'wallet'})
        keep.append('wallet')
    df = df[keep].copy()

    df['date']   = pd.to_datetime(df['date'], errors='coerce', dayfirst=True)
    df['amount'] = df['amount'].apply(_clean_number)
    df = df.dropna(subset=['date', 'amount', 'article'])
    df = df[df['article'].str.strip().ne('')]
    if 'wallet' in df.columns:
        df = df[~df['wallet'].isin(BANK_WALLETS)].drop(columns=['wallet'])
    return df


def _get_dds_gs(gc):
    """Читает лист ДДС из Google Sheets как сырую матрицу строк (аналог header=None)."""
    try:
        sh = gc.open_by_key(GSHEET_ID)
        ws = sh.worksheet('ДДС')
        data = ws.get_all_values()
    except Exception as e:
        raise FileNotFoundError(f'Не удалось загрузить лист ДДС из Google Sheets: {e}')
    return pd.DataFrame(data)


def get_current_file_info():
    if USE_GSHEETS and CREDENTIALS_FILE.exists():
        return {
            'path':   f'https://docs.google.com/spreadsheets/d/{GSHEET_ID}',
            'name':   'Google Sheets (авто)',
            'exists': True,
            'source': 'gsheets',
        }
    if XLSX_IN is None:
        return {'path': '', 'name': 'файл не выбран', 'exists': False, 'source': 'none'}
    return {
        'path':   str(XLSX_IN),
        'name':   XLSX_IN.name,
        'exists': XLSX_IN.exists(),
        'source': 'xlsx',
    }


def _load_sheet(name, dc, ac, rc, hr, wc=None):
    try:
        df = pd.read_excel(XLSX_IN, sheet_name=name, header=hr, dtype=str)
    except Exception:
        return pd.DataFrame()
    cols = df.columns.tolist()
    if max(dc, ac, rc) >= len(cols):
        return pd.DataFrame()
    df = df.rename(columns={cols[dc]: 'date', cols[ac]: 'amount', cols[rc]: 'article'})
    keep = ['date', 'amount', 'article']
    if wc is not None and wc < len(cols):
        df = df.rename(columns={cols[wc]: 'wallet'}); keep.append('wallet')
    df = df[keep].copy()
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
    df = df.dropna(subset=['date', 'amount', 'article'])
    df = df[df['article'].str.strip().ne('')]
    if 'wallet' in df.columns:
        df = df[~df['wallet'].isin(BANK_WALLETS)].drop(columns=['wallet'])
    return df


def get_tx():
    global _TX
    if _TX is None:
        if USE_GSHEETS and CREDENTIALS_FILE.exists():
            print("Загрузка транзакций из Google Sheets...")
            gc = _get_gs_client()
            frames = [_load_sheet_gs(gc, *c, wc=1) for c in JOURNAL_SHEETS] + \
                     [_load_sheet_gs(gc, *c)       for c in BANK_SHEETS]
        else:
            if XLSX_IN is None:
                raise FileNotFoundError('Источник данных не выбран. Загрузите .xlsx файл.')
            if not XLSX_IN.exists():
                raise FileNotFoundError(f'Файл не найден: {XLSX_IN}')
            print("Загрузка транзакций из Excel...")
            frames = [_load_sheet(*c, wc=1) for c in JOURNAL_SHEETS] + \
                     [_load_sheet(*c)       for c in BANK_SHEETS]
        _TX = pd.concat([f for f in frames if not f.empty], ignore_index=True)
        _TX['article'] = _TX['article'].str.strip().replace(RENAME_ARTICLES)
        print(f"  Транзакций: {len(_TX):,}")
    return _TX


def get_dds():
    """Кэшированное чтение листа ДДС (числа, без заголовков)."""
    global _DDS
    if _DDS is None:
        if USE_GSHEETS and CREDENTIALS_FILE.exists():
            print("Загрузка листа ДДС из Google Sheets...")
            gc = _get_gs_client()
            _DDS = _get_dds_gs(gc)
        else:
            if XLSX_IN is None:
                raise FileNotFoundError('Источник данных не выбран. Загрузите .xlsx файл.')
            if not XLSX_IN.exists():
                raise FileNotFoundError(f'Файл не найден: {XLSX_IN}')
            print("Загрузка листа ДДС...")
            _DDS = pd.read_excel(XLSX_IN, sheet_name='ДДС', header=None)
    return _DDS


# ── Revenue from ДДС sheet ─────────────────────────────────────────────────────
def get_revenue(year, months):
    """
    Выручка из строки 18 листа ДДС
    «Поступления от клиентов (продажа товаров/услуг)»
    Суммируем значения по указанным месяцам для данного года.
    """
    dds = get_dds()
    offset = DDS_YEAR_OFFSETS.get(year)
    if offset is None:
        return 0.0
    total = 0.0
    for m in months:
        col_idx = offset + m
        if col_idx < dds.shape[1]:
            v = dds.iloc[DDS_REVENUE_ROW, col_idx]
            try:
                total += _clean_number(v)
            except (ValueError, TypeError):
                pass
    return total


# ── Core analytics ─────────────────────────────────────────────────────────────
def is_credit_article(name):
    return name in CREDIT_ARTICLES_SET


def build_pivot(tx, date_from, date_to, months, credit=False):
    """
    credit=False → обычные расходы (без кредитов):
                   amount < 0 → расход (val > 0)
                   amount > 0 → возврат/корректировка (val < 0, вычитается из итога)
    credit=True  → все кредитные операции (amount < 0 И amount > 0)
    """
    # date_to is inclusive to end of day (handles Excel dates with time components)
    date_to_eod = pd.Timestamp(date_to) + pd.Timedelta(hours=23, minutes=59, seconds=59)
    base = tx[(tx['date'] >= date_from) & (tx['date'] <= date_to_eod) &
              (~tx['article'].isin(EXCLUDE_ARTICLES))].copy()
    # Filter to exact months (handles non-contiguous month selection)
    base = base[base['date'].dt.month.isin(months)]

    if credit:
        base = base[base['article'].apply(is_credit_article)]
        # Для кредитов: сохраняем знак (выдача = расход < 0, возврат = доход > 0)
        base['val'] = base['amount']
    else:
        # Берём все НЕ-кредитные и НЕ-выручковые статьи
        # val = -amount: расход (-5000) → val=5000
        mask = (~base['article'].apply(is_credit_article) &
                ~base['article'].apply(is_revenue_article))
        base = base[mask]
        base['val'] = -base['amount']

    base['mon'] = base['date'].dt.month
    pv = base.groupby(['article', 'mon'])['val'].sum().unstack(fill_value=0)
    for m in months:
        if m not in pv.columns:
            pv[m] = 0
    return pv[months]


def detect_groups(articles):
    pfx = defaultdict(list)
    for a in articles:
        pfx[a.split()[0].rstrip('/')].append(a)
    out = {}
    for k, arts in pfx.items():
        # Никогда не объединять статьи разных видов расходов в одну группу
        by_cat = defaultdict(list)
        for a in arts:
            by_cat[get_expense_category(a)].append(a)
        multi_cat = len(by_cat) > 1
        for cat, cat_arts in by_cat.items():
            if len(cat_arts) >= 2:
                # Добавляем суффикс только если одно и то же первое слово используется
                # в обеих категориях (чтобы не было дублей ключей в gmap)
                suffix = f' ({"произв." if cat == "Производственные" else "адм."})' if multi_cat else ''
                lbl = k.upper() + suffix
            else:
                lbl = cat_arts[0]
            for a in cat_arts:
                out[a] = lbl
    return out


def get_date_range(period_type, period_value, year):
    y = int(year)
    if period_type == 'month':
        m = int(period_value)
        last = calendar.monthrange(y, m)[1]
        return f'{y}-{m:02d}-01', f'{y}-{m:02d}-{last:02d}', [m]
    elif period_type == 'quarter':
        q = int(period_value)
        ms = {1: [1, 2, 3], 2: [4, 5, 6], 3: [7, 8, 9], 4: [10, 11, 12]}[q]
        last = calendar.monthrange(y, ms[-1])[1]
        return f'{y}-{ms[0]:02d}-01', f'{y}-{ms[-1]:02d}-{last:02d}', ms
    elif period_type == 'months':
        # period_value = "1,3,5" — произвольный набор месяцев
        ms = sorted(set(int(x) for x in period_value.split(',') if x.strip()))
        if not ms:
            ms = [1]
        last = calendar.monthrange(y, ms[-1])[1]
        return f'{y}-{ms[0]:02d}-01', f'{y}-{ms[-1]:02d}-{last:02d}', ms
    elif period_type == 'range':
        # period_value = "MM-DD|MM-DD" — произвольный диапазон дней
        from_md, to_md = period_value.split('|')
        date_from = f'{y}-{from_md}'
        date_to = f'{y}-{to_md}'
        from_m = int(from_md.split('-')[0])
        to_m = int(to_md.split('-')[0])
        ms = list(range(from_m, to_m + 1)) if from_m <= to_m else [from_m]
        return date_from, date_to, ms
    else:
        return f'{y}-01-01', f'{y}-12-31', list(range(1, 13))


def _build_rows(pv_new, pv_old, months_new, months_old, rev_old, rev_new, add_cat_headers=True):
    """Построить список строк из двух пивотов. Возвращает (rows, total_old, total_new)."""
    all_arts = list(dict.fromkeys(
        list(pv_new.index) + [a for a in pv_old.index if a not in pv_new.index]
    ))
    if not all_arts:
        return [], 0, 0

    art_to_grp = detect_groups(all_arts)
    gmap = defaultdict(list)
    for a, g in art_to_grp.items():
        gmap[g].append(a)
    # Сортировка: Производственные → Административные,
    # внутри — алфавит по возрастанию; дивиденды/ЕНП — в конце административных
    order = sorted(gmap, key=lambda g: (
        0 if _get_grp_category(gmap[g]) == 'Производственные' else 1,
        1 if (_get_grp_category(gmap[g]) == 'Административные' and _is_admin_tail(g)) else 0,
        g.lower(),
    ))

    def safe_pct(v, base):
        return round(v / base * 100, 1) if base else None

    rows = []
    g_old_total = g_new_total = 0
    prev_cat = None

    for grp in order:
        arts = gmap[grp]
        cat = _get_grp_category(arts)

        # Вставить разделитель-заголовок при смене категории
        if add_cat_headers and cat != prev_cat:
            lbl = 'ПРОИЗВОДСТВЕННЫЕ РАСХОДЫ' if cat == 'Производственные' else 'АДМИНИСТРАТИВНЫЕ РАСХОДЫ'
            rows.append({'type': 'cat_head', 'label': lbl, 'category': cat})
            prev_cat = cat

        g_new_vals = [float(pv_new.loc[pv_new.index.isin(arts), m].sum()) if m in pv_new.columns else 0.0 for m in months_new]
        g_old_vals = [float(pv_old.loc[pv_old.index.isin(arts), m].sum()) if m in pv_old.columns else 0.0 for m in months_old]
        g_new_sum = sum(g_new_vals)
        g_old_sum = sum(g_old_vals)
        g_old_total += g_old_sum
        g_new_total += g_new_sum

        delta = g_new_sum - g_old_sum
        pct = round(delta / g_old_sum * 100, 1) if g_old_sum else (100 if g_new_sum else 0)
        pct_ro = safe_pct(g_old_sum, rev_old)
        pct_rn = safe_pct(g_new_sum, rev_new)
        dpp = round(pct_rn - pct_ro, 1) if pct_ro is not None and pct_rn is not None else None

        if g_old_sum == 0 and g_new_sum == 0:    assess = '—'
        elif g_old_sum == 0:                      assess = 'новая статья'
        elif g_new_sum == 0:                      assess = 'статья закрыта'
        elif pct > 20:                            assess = f'рост {abs(pct):.0f}%'
        elif pct < -20:                           assess = f'снижение {abs(pct):.0f}%'
        else:                                     assess = 'стабильно'

        rows.append({
            'type': 'group', 'label': grp, 'single': len(arts) == 1,
            'category': cat,
            'months_new': [round(v) for v in g_new_vals],
            'months_old': [round(v) for v in g_old_vals],
            'avg_new': round(g_new_sum / len(months_new)) if months_new else 0,
            'total_new': round(g_new_sum), 'total_old': round(g_old_sum),
            'delta': round(delta), 'pct': pct, 'assess': assess,
            'pct_rev_old': pct_ro, 'pct_rev_new': pct_rn, 'delta_pp': dpp,
        })

        if len(arts) > 1:
            for art in sorted(arts,
                    key=lambda a: abs(pv_new.loc[a, months_new].sum()) if a in pv_new.index else 0,
                    reverse=True):
                a_new = [float(pv_new.loc[art, m]) if art in pv_new.index and m in pv_new.columns else 0 for m in months_new]
                a_old = [float(pv_old.loc[art, m]) if art in pv_old.index and m in pv_old.columns else 0 for m in months_old]
                an = sum(a_new); ao = sum(a_old); d = an - ao
                p = round(d / ao * 100, 1) if ao else (100 if an else 0)
                rows.append({
                    'type': 'detail', 'label': art,
                    'category': cat,
                    'months_new': [round(v) for v in a_new],
                    'months_old': [round(v) for v in a_old],
                    'avg_new': round(an / len(months_new)) if months_new else 0,
                    'total_new': round(an), 'total_old': round(ao),
                    'delta': round(d), 'pct': p, 'assess': '',
                    'pct_rev_old': safe_pct(ao, rev_old),
                    'pct_rev_new': safe_pct(an, rev_new),
                    'delta_pp': None,
                })

    return rows, g_old_total, g_new_total


def _period_label(period_type, period_value, year):
    y = int(year)
    if period_type == 'month':
        return f"{MONTH_RU[int(period_value)]} {y}"
    elif period_type == 'quarter':
        return f"Q{period_value} {y}"
    elif period_type == 'months':
        ms_sorted = sorted(int(x) for x in period_value.split(',') if x.strip())
        return ', '.join(MONTH_RU[m][:3] for m in ms_sorted) + f" {y}"
    elif period_type == 'range':
        from_md, to_md = period_value.split('|')
        return f"{from_md}–{to_md} {y}"
    else:
        return str(y)


def get_revenue_breakdown(tx, date_from, date_to, months):
    """Разбивка выручки по статьям (для вкладки сравнения)."""
    date_to_eod = pd.Timestamp(date_to) + pd.Timedelta(hours=23, minutes=59, seconds=59)
    base = tx[(tx['date'] >= date_from) & (tx['date'] <= date_to_eod) &
              (~tx['article'].isin(EXCLUDE_ARTICLES)) &
              (tx['article'].apply(is_revenue_article))].copy()
    base = base[base['date'].dt.month.isin(months)]
    if base.empty:
        return []
    base['val'] = base.apply(lambda r: _rev_val(r['article'], r['amount']), axis=1)
    base['mon'] = base['date'].dt.month
    pv = base.groupby(['article', 'mon'])['val'].sum().unstack(fill_value=0)
    for m in months:
        if m not in pv.columns:
            pv[m] = 0
    pv = pv[months]
    result = []
    for art in pv.index:
        vals = [round(float(pv.loc[art, m])) for m in months]
        total = round(sum(vals))
        result.append({'label': art, 'months': vals, 'total': total})
    result.sort(key=lambda x: -x['total'])
    return result


def get_interval_revenue_breakdown(tx, date_from, date_to, intervals, iv_fn):
    """Разбивка выручки по статьям (для вкладки анализа за период)."""
    date_to_eod = pd.Timestamp(date_to) + pd.Timedelta(hours=23, minutes=59, seconds=59)
    base = tx[(tx['date'] >= date_from) & (tx['date'] <= date_to_eod) &
              (~tx['article'].isin(EXCLUDE_ARTICLES)) &
              (tx['article'].apply(is_revenue_article))].copy()
    if base.empty:
        return []
    base['val'] = base.apply(lambda r: _rev_val(r['article'], r['amount']), axis=1)
    base['iv'] = base['date'].apply(iv_fn)
    pv = base.groupby(['article', 'iv'])['val'].sum().unstack(fill_value=0)
    for iv in intervals:
        if iv not in pv.columns:
            pv[iv] = 0
    pv = pv[intervals]
    result = []
    for art in pv.index:
        vals = [round(float(pv.loc[art, iv])) for iv in intervals]
        total = round(sum(vals))
        result.append({
            'label': art,
            'iv_values': vals,
            'total': total,
            'avg': round(total / max(len(intervals), 1)),
        })
    result.sort(key=lambda x: -x['total'])
    return result


def build_report(period_type, period_value_new, year_new, period_value_old, year_old):
    tx = get_tx()
    yn, yo = int(year_new), int(year_old)
    new_from, new_to, months_new = get_date_range(period_type, period_value_new, yn)
    old_from, old_to, months_old = get_date_range(period_type, period_value_old, yo)

    # Выручка из листа ДДС строка 18
    rev_new = get_revenue(yn, months_new)
    rev_old = get_revenue(yo, months_old)

    # ── Расходы (без кредитов) ────────────────────────────────────────────────
    pv_new = build_pivot(tx, new_from, new_to, months_new, credit=False)
    pv_old = build_pivot(tx, old_from, old_to, months_old, credit=False)
    pv_new = pv_new[(pv_new > 0).any(axis=1)]
    pv_old = pv_old[(pv_old > 0).any(axis=1)]

    expense_rows, exp_old, exp_new = _build_rows(pv_new, pv_old, months_new, months_old, rev_old, rev_new)

    # Grand total расходы
    gd = exp_new - exp_old
    gp = round(gd / exp_old * 100, 1) if exp_old else 0
    def sp(v, b): return round(v / b * 100, 1) if b else None
    expense_rows.append({
        'type': 'grand', 'label': 'ИТОГО РАСХОДЫ',
        'months_new': [], 'months_old': [],
        'avg_new': round(exp_new / len(months_new)) if months_new else 0,
        'total_new': round(exp_new), 'total_old': round(exp_old),
        'delta': round(gd), 'pct': gp, 'assess': '',
        'pct_rev_old': sp(exp_old, rev_old), 'pct_rev_new': sp(exp_new, rev_new),
        'delta_pp': round((sp(exp_new, rev_new) or 0) - (sp(exp_old, rev_old) or 0), 1)
                   if rev_old and rev_new else None,
    })

    # ── Кредиты и займы (отдельный блок) ──────────────────────────────────────
    cr_new = build_pivot(tx, new_from, new_to, months_new, credit=True)
    cr_old = build_pivot(tx, old_from, old_to, months_old, credit=True)
    credit_rows, cr_old_t, cr_new_t = _build_rows(cr_new, cr_old, months_new, months_old, rev_old, rev_new, add_cat_headers=False)

    if credit_rows:
        crd = cr_new_t - cr_old_t
        credit_rows.append({
            'type': 'grand', 'label': 'ИТОГО КРЕДИТЫ И ЗАЙМЫ',
            'months_new': [], 'months_old': [],
            'avg_new': round(cr_new_t / len(months_new)) if months_new else 0,
            'total_new': round(cr_new_t), 'total_old': round(cr_old_t),
            'delta': round(crd),
            'pct': round(crd / cr_old_t * 100, 1) if cr_old_t else 0,
            'assess': '', 'pct_rev_old': None, 'pct_rev_new': None, 'delta_pp': None,
        })

    plabel_new = _period_label(period_type, period_value_new, yn)
    plabel_old = _period_label(period_type, period_value_old, yo)

    # Помесячная выручка (новый период, для вкладки Данные)
    rev_new_months = [round(get_revenue(yn, [m])) for m in months_new]
    rev_old_months = [round(get_revenue(yo, [m])) for m in months_old]

    rev_delta = round(rev_new) - round(rev_old)
    rev_pct = round(rev_delta / rev_old * 100, 1) if rev_old else 0

    # Разбивка выручки по статьям из транзакций
    rev_breakdown_new = get_revenue_breakdown(tx, new_from, new_to, months_new)
    rev_breakdown_old = get_revenue_breakdown(tx, old_from, old_to, months_old)

    return {
        'period_label': plabel_new,
        'period_label_new': plabel_new,
        'period_label_old': plabel_old,
        'period_type': period_type,
        'period_value_new': period_value_new,
        'period_value_old': period_value_old,
        'year_new': yn, 'year_old': yo,
        'months': months_new,
        'month_names': [MONTH_RU[m] for m in months_new],
        'rev_new': round(rev_new),
        'rev_old': round(rev_old),
        'rev_new_months': rev_new_months,
        'rev_old_months': rev_old_months,
        'rev_delta': rev_delta,
        'rev_pct': rev_pct,
        'rev_breakdown_new': rev_breakdown_new,
        'rev_breakdown_old': rev_breakdown_old,
        'rows': expense_rows,
        'credit_rows': credit_rows,
    }


# ── XLSX generation ────────────────────────────────────────────────────────────
def _fill(h): return PatternFill('solid', fgColor=h)
def _fnt(bold=False, color='000000', size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
def _aln(h='right', v='center'): return Alignment(horizontal=h, vertical=v)
_s = Side(style='thin', color='B0BEC5')
_BRD = Border(left=_s, right=_s, top=_s, bottom=_s)
_NUM = '#,##0'


def _wcell(ws, r, c, val, fl, fn, fmt=None, ha='right'):
    cell = ws.cell(r, c, val)
    cell.fill = fl; cell.font = fn; cell.border = _BRD; cell.alignment = _aln(ha)
    if fmt: cell.number_format = fmt
    return cell


def _row_style(row):
    t = row['type']
    if t == 'grand':
        return _fill('0D1E35'), _fnt(bold=True, color='FFFFFF', size=11)
    if t == 'detail':
        return _fill('FAFBFD'), _fnt(italic=True, color='444466', size=9)
    if row.get('single'):
        return _fill('CFD8E8'), _fnt(bold=True, color='1A1A2E', size=10)
    return _fill('1A3A5C'), _fnt(bold=True, color='FFFFFF', size=10)


def _credit_style(row):
    t = row['type']
    if t == 'grand':
        return _fill('3d2c1e'), _fnt(bold=True, color='FFFFFF', size=11)
    if t == 'detail':
        return _fill('FFF8F0'), _fnt(italic=True, color='7d5a3a', size=9)
    if row.get('single'):
        return _fill('F5E6D0'), _fnt(bold=True, color='5a3a1a', size=10)
    return _fill('8B5E3C'), _fnt(bold=True, color='FFFFFF', size=10)


def _write_block(ws, start_row, all_rows, months, mnames, yn, style_fn):
    """Write a block of rows to a data sheet. Returns next row."""
    n = len(months); r = start_row
    last_col = get_column_letter(n + 3)
    for row in all_rows:
        if row.get('type') == 'cat_head':
            cat = row.get('category', '')
            cat_fl = _fill('1B5E3A') if cat == 'Производственные' else _fill('2C4A7C')
            ws.merge_cells(f'A{r}:{last_col}{r}')
            cell = ws.cell(r, 1, row.get('label', ''))
            cell.fill = cat_fl; cell.font = _fnt(bold=True, color='FFFFFF', size=10)
            cell.alignment = _aln('center'); cell.border = _BRD
            r += 1
            continue
        fl, fn = style_fn(row)
        label = ('\u00a0' * 4 + 'в т.ч. ' + row['label']) if row['type'] == 'detail' else row['label']
        _wcell(ws, r, 1, label, fl, fn, ha='left')
        for i, v in enumerate(row.get('months_new', [])):
            _wcell(ws, r, i + 2, v, fl, fn, _NUM)
        if not row.get('months_new'):
            for i in range(n):
                _wcell(ws, r, i + 2, '', fl, fn)
        _wcell(ws, r, n + 2, row.get('avg_new', ''), fl, fn, _NUM)
        _wcell(ws, r, n + 3, row.get('total_new', ''), fl, fn, _NUM)
        r += 1
    return r


def _write_cmp_block(ws, start_row, all_rows, style_fn):
    """Write a block of rows to a comparison sheet. Returns next row."""
    r = start_row
    for row in all_rows:
        if row.get('type') == 'cat_head':
            cat = row.get('category', '')
            cat_fl = _fill('1B5E3A') if cat == 'Производственные' else _fill('2C4A7C')
            ws.merge_cells(f'A{r}:I{r}')
            cell = ws.cell(r, 1, row.get('label', ''))
            cell.fill = cat_fl; cell.font = _fnt(bold=True, color='FFFFFF', size=10)
            cell.alignment = _aln('center'); cell.border = _BRD
            r += 1
            continue
        fl, fn = style_fn(row)
        label = ('\u00a0' * 4 + 'в т.ч. ' + row['label']) if row['type'] == 'detail' else row['label']
        _wcell(ws, r, 1, label, fl, fn, ha='left')
        _wcell(ws, r, 2, row.get('total_old', ''), fl, fn, _NUM)
        _wcell(ws, r, 3, row.get('total_new', ''), fl, fn, _NUM)
        _wcell(ws, r, 4, row.get('delta', ''), fl, fn, '+#,##0;-#,##0;—')
        pct = row.get('pct')
        _wcell(ws, r, 5, f"{pct:+.1f}%" if pct is not None else '—', fl, fn)
        _wcell(ws, r, 6, row.get('assess', ''), fl, fn, ha='center')
        pro = row.get('pct_rev_old')
        prn = row.get('pct_rev_new')
        _wcell(ws, r, 7, f"{pro:.1f}%" if pro is not None else '—', fl, fn)
        _wcell(ws, r, 8, f"{prn:.1f}%" if prn is not None else '—', fl, fn)
        dpp = row.get('delta_pp')
        _wcell(ws, r, 9, f"{dpp:+.1f} пп" if dpp is not None else '—', fl, fn, ha='center')
        r += 1
    return r


def generate_xlsx(report):
    wb = openpyxl.Workbook()
    rows = report['rows']
    credit = report.get('credit_rows', [])
    months = report['months']; mnames = report['month_names']
    yn = report['year_new']; yo = report['year_old']
    n = len(months)

    # ── Sheet 1: Data ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = (report.get('period_label_new') or report['period_label'])[:31]
    ws1.column_dimensions['A'].width = 50
    for i in range(1, n + 3):
        ws1.column_dimensions[chr(ord('A') + i)].width = 15

    last_col = chr(ord('A') + n + 2)
    ws1.merge_cells(f'A1:{last_col}1')
    plbl_new = report.get('period_label_new') or report['period_label']
    plbl_old = report.get('period_label_old') or report['period_label']
    ws1['A1'].value = f"Расходы — {plbl_new}"
    ws1['A1'].font = _fnt(bold=True, color='FFFFFF', size=12)
    ws1['A1'].fill = _fill('1A3A5C'); ws1['A1'].alignment = _aln('center')

    hdrs = ['Статья / Группа'] + [f'{m} {yn}' for m in mnames] + ['Среднее/мес', 'Итого']
    for ci, h in enumerate(hdrs, 1):
        c = ws1.cell(2, ci, h)
        c.font = _fnt(bold=True, color='FFFFFF', size=9)
        c.fill = _fill('1A3A5C'); c.border = _BRD
        c.alignment = _aln('center' if ci == 1 else 'right')

    r = _write_block(ws1, 3, rows, months, mnames, yn, _row_style)

    if credit:
        r += 1  # blank row
        sec = ws1.cell(r, 1, 'КРЕДИТЫ И ЗАЙМЫ')
        sec.font = _fnt(bold=True, color='8B5E3C', size=11)
        sec.fill = _fill('F5E6D0'); sec.border = _BRD
        ws1.merge_cells(f'A{r}:{last_col}{r}')
        sec.alignment = _aln('center')
        r += 1
        _write_block(ws1, r, credit, months, mnames, yn, _credit_style)

    ws1.freeze_panes = 'B3'

    # ── Sheet 2: Comparison ─────────────────────────────────────────────────
    ws2 = wb.create_sheet(('vs ' + plbl_old)[:31])
    ws2.column_dimensions['A'].width = 50
    for col in list('BCDEFGHI'):
        ws2.column_dimensions[col].width = 15

    ws2.merge_cells('A1:I1')
    ws2['A1'].value = (f"Сравнение: {plbl_new} vs {plbl_old}   "
                       f"Выручка {plbl_old}: {report['rev_old']:,} | "
                       f"Выручка {plbl_new}: {report['rev_new']:,}")
    ws2['A1'].font = _fnt(bold=True, color='FFFFFF', size=11)
    ws2['A1'].fill = _fill('1A3A5C'); ws2['A1'].alignment = _aln('center')

    hdrs2 = ['Статья / Группа', plbl_old, plbl_new,
             'Изменение', 'Тренд %', 'Оценка',
             f'% выр. {plbl_old}', f'% выр. {plbl_new}', 'Д пп']
    for ci, h in enumerate(hdrs2, 1):
        c = ws2.cell(2, ci, h)
        c.font = _fnt(bold=True, color='FFFFFF', size=9)
        c.fill = _fill('1A3A5C'); c.border = _BRD
        c.alignment = _aln('center' if ci == 1 else 'right')

    r = _write_cmp_block(ws2, 3, rows, _row_style)

    if credit:
        r += 1
        sec = ws2.cell(r, 1, 'КРЕДИТЫ И ЗАЙМЫ')
        sec.font = _fnt(bold=True, color='8B5E3C', size=11)
        sec.fill = _fill('F5E6D0'); sec.border = _BRD
        ws2.merge_cells(f'A{r}:I{r}')
        sec.alignment = _aln('center')
        r += 1
        _write_cmp_block(ws2, r, credit, _credit_style)

    ws2.freeze_panes = 'B3'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ── XLSX for period analysis ───────────────────────────────────────────────────
def generate_xlsx_period(report):
    wb = openpyxl.Workbook()
    rows   = report['rows']
    credit = report.get('credit_rows', [])
    ivs    = report['intervals']
    lbls   = report['interval_labels']
    n      = len(ivs)
    rev_row   = report.get('rev_row', {})
    exp_total = report.get('exp_total', 0) or 1  # avoid div-by-zero

    # Columns: label | category | iv1…ivN | Среднее | Итого | % выручки
    TOTAL_COLS = n + 5
    last_col   = get_column_letter(TOTAL_COLS)

    ws = wb.active
    d_from = report.get('date_from', '')
    d_to   = report.get('date_to',   '')
    ws.title = f'Анализ {d_from}—{d_to}'[:31]

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 14
    for ci in range(3, TOTAL_COLS + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # Title
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'].value = f'Анализ расходов: {d_from} — {d_to}'
    ws['A1'].font  = _fnt(bold=True, color='FFFFFF', size=12)
    ws['A1'].fill  = _fill('1A3A5C')
    ws['A1'].alignment = _aln('center')

    # Header
    hdrs = ['Статья / Группа', 'Вид расходов'] + lbls + ['Среднее', 'Итого', '% выручки']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(2, ci, h)
        c.font = _fnt(bold=True, color='FFFFFF', size=9)
        c.fill = _fill('1A3A5C'); c.border = _BRD
        c.alignment = _aln('left' if ci == 1 else 'center')

    def _write_pa_row(r_idx, row, is_credit=False):
        t = row.get('type', 'group')
        if t == 'cat_head':
            cat   = row.get('category', '')
            fl_c  = _fill('1B5E3A') if cat == 'Производственные' else _fill('2C4A7C')
            ws.merge_cells(f'A{r_idx}:{last_col}{r_idx}')
            cell  = ws.cell(r_idx, 1, row.get('label', ''))
            cell.fill = fl_c; cell.font = _fnt(bold=True, color='FFFFFF', size=10)
            cell.alignment = _aln('center'); cell.border = _BRD
            return r_idx + 1

        if is_credit:
            if t == 'grand':   fl, fn = _fill('3d2c1e'), _fnt(bold=True, color='FFFFFF', size=11)
            elif t == 'detail':fl, fn = _fill('FFF8F0'), _fnt(italic=True, color='7d5a3a', size=9)
            elif row.get('single'): fl, fn = _fill('F5E6D0'), _fnt(bold=True, color='5a3a1a', size=10)
            else:              fl, fn = _fill('8B5E3C'), _fnt(bold=True, color='FFFFFF', size=10)
        else:
            fl, fn = _row_style(row)

        indent = '\u00a0' * 4 + 'в т.ч. ' if t == 'detail' else ''
        _wcell(ws, r_idx, 1, indent + row.get('label', ''), fl, fn, ha='left')
        _wcell(ws, r_idx, 2, row.get('category', ''), fl, fn, ha='center')
        ivv = row.get('iv_values', [])
        for i, v in enumerate(ivv, 3):
            _wcell(ws, r_idx, i, v, fl, fn, _NUM)
        if not ivv:
            for i in range(3, n + 3):
                _wcell(ws, r_idx, i, '', fl, fn)
        _wcell(ws, r_idx, n + 3, row.get('avg',   ''), fl, fn, _NUM)
        _wcell(ws, r_idx, n + 4, row.get('total', ''), fl, fn, _NUM)
        pct_rev = row.get('pct_rev')
        _wcell(ws, r_idx, n + 5,
               f"{pct_rev:.1f}%" if pct_rev is not None else '—', fl, fn)
        return r_idx + 1

    # Revenue row
    r = 3
    rev_fl = _fill('1B7A3D'); rev_fn = _fnt(bold=True, color='FFFFFF', size=11)
    _wcell(ws, r, 1, 'ВЫРУЧКА', rev_fl, rev_fn, ha='left')
    _wcell(ws, r, 2, '',        rev_fl, rev_fn)
    for i, v in enumerate(rev_row.get('iv_values', []), 3):
        _wcell(ws, r, i, v, rev_fl, rev_fn, _NUM)
    if not rev_row.get('iv_values'):
        for i in range(3, n + 3): _wcell(ws, r, i, '', rev_fl, rev_fn)
    _wcell(ws, r, n + 3, rev_row.get('avg',   ''), rev_fl, rev_fn, _NUM)
    _wcell(ws, r, n + 4, rev_row.get('total', ''), rev_fl, rev_fn, _NUM)
    _wcell(ws, r, n + 5, '—', rev_fl, rev_fn)
    r += 1

    for row in rows:
        r = _write_pa_row(r, row)

    if credit:
        r += 1
        sec = ws.cell(r, 1, 'КРЕДИТЫ И ЗАЙМЫ')
        sec.font = _fnt(bold=True, color='8B5E3C', size=11)
        sec.fill = _fill('F5E6D0'); sec.border = _BRD
        ws.merge_cells(f'A{r}:{last_col}{r}')
        sec.alignment = _aln('center')
        r += 1
        for row in credit:
            r = _write_pa_row(r, row, is_credit=True)

    ws.freeze_panes = 'C3'
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html', file_info=get_current_file_info())


@app.route('/health', methods=['GET'])
def health():
    return jsonify({
        'ok': True,
        'service': 'dds_allada',
        'file': get_current_file_info(),
    })


@app.route('/api/file-info', methods=['GET'])
def api_file_info():
    return jsonify(get_current_file_info())


@app.route('/api/upload', methods=['POST'])
def api_upload():
    global XLSX_IN
    try:
        if 'file' not in request.files:
            return jsonify({'ok': False, 'error': 'Файл не передан.'}), 400

        uploaded = request.files['file']
        if not uploaded or not uploaded.filename:
            return jsonify({'ok': False, 'error': 'Выберите файл для загрузки.'}), 400

        safe_name = secure_filename(uploaded.filename)
        ext = Path(safe_name).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            return jsonify({'ok': False, 'error': 'Разрешен только формат .xlsx.'}), 400

        upload_dir = Path(__file__).resolve().parent / 'uploads'
        upload_dir.mkdir(parents=True, exist_ok=True)
        target = upload_dir / TARGET_FILE_NAME
        uploaded.save(target)

        # Проверяем, что файл действительно читается как Excel workbook.
        openpyxl.load_workbook(target, read_only=True).close()

        XLSX_IN = target
        reset_cache()

        return jsonify({
            'ok': True,
            'message': f'Файл {safe_name} успешно загружен.',
            'file_info': get_current_file_info(),
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Ошибка обработки файла: {e}'}), 400


@app.route('/api/report', methods=['POST'])
def api_report():
    data = request.json
    try:
        # Support both old single period_value and new separate period_value_new/old
        pv_new = data.get('period_value_new') or data.get('period_value', '1')
        pv_old = data.get('period_value_old') or data.get('period_value', '1')
        report = build_report(
            data['period_type'],
            pv_new,
            data['year_new'],
            pv_old,
            data['year_old'],
        )
        return jsonify(report)
    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/download', methods=['POST'])
def api_download():
    data = request.json
    try:
        pv_new = data.get('period_value_new') or data.get('period_value', '1')
        pv_old = data.get('period_value_old') or data.get('period_value', '1')
        report = build_report(
            data['period_type'],
            pv_new,
            data['year_new'],
            pv_old,
            data['year_old'],
        )
        buf = generate_xlsx(report)
        lbl_new = report['period_label_new'].replace(' ', '_').replace(',', '').replace('–', '-')
        lbl_old = report['period_label_old'].replace(' ', '_').replace(',', '').replace('–', '-')
        fname = f"comparison_{lbl_new}_vs_{lbl_old}.xlsx"
        return send_file(buf, download_name=fname, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 400


# ── Period analysis (no comparison) ───────────────────────────────────────────

def _make_interval_col(interval, from_dt, to_dt):
    """Returns (intervals, labels, date→key lambda)."""
    if interval == 'year':
        years = list(range(from_dt.year, to_dt.year + 1))
        ivs   = [str(y) for y in years]
        lbls  = ivs.copy()
        fn    = lambda d: str(d.year)
    elif interval == 'month':
        ivs, lbls = [], []
        curr = from_dt.replace(day=1)
        while curr <= to_dt:
            ivs.append(f'{curr.year}-{curr.month:02d}')
            lbls.append(f'{MONTH_RU[curr.month]} {curr.year}')
            curr = curr + pd.DateOffset(months=1)
        fn = lambda d: f'{d.year}-{d.month:02d}'
    else:  # day
        days = pd.date_range(from_dt, to_dt, freq='D')
        ivs  = [d.strftime('%Y-%m-%d') for d in days]
        lbls = [d.strftime('%d.%m') for d in days]
        fn   = lambda d: d.strftime('%Y-%m-%d')
    return ivs, lbls, fn


def _build_interval_pivot(tx, date_from, date_to, intervals, iv_fn, credit=False):
    date_to_eod = pd.Timestamp(date_to) + pd.Timedelta(hours=23, minutes=59, seconds=59)
    base = tx[(tx['date'] >= date_from) & (tx['date'] <= date_to_eod) &
              (~tx['article'].isin(EXCLUDE_ARTICLES))].copy()
    if credit:
        base = base[base['article'].apply(is_credit_article)]
        base['val'] = base['amount']
    else:
        # НЕ-кредитные и НЕ-выручковые статьи; val = -amount
        mask = (~base['article'].apply(is_credit_article) &
                ~base['article'].apply(is_revenue_article))
        base = base[mask]
        base['val'] = -base['amount']
    if base.empty:
        return pd.DataFrame(0, index=pd.Index([], name='article'), columns=intervals)
    base['iv'] = base['date'].apply(iv_fn)
    pv = base.groupby(['article', 'iv'])['val'].sum().unstack(fill_value=0)
    for iv in intervals:
        if iv not in pv.columns:
            pv[iv] = 0
    return pv[intervals]


def _build_interval_rows(pv, intervals, rev_total, add_cat_headers=True):
    all_arts = list(pv.index)
    if not all_arts:
        return [], 0

    art_to_grp = detect_groups(all_arts)
    gmap = defaultdict(list)
    for a, g in art_to_grp.items():
        gmap[g].append(a)
    order = sorted(gmap, key=lambda g: (
        0 if _get_grp_category(gmap[g]) == 'Производственные' else 1,
        1 if (_get_grp_category(gmap[g]) == 'Административные' and _is_admin_tail(g)) else 0,
        g.lower(),
    ))

    def sp(v): return round(v / rev_total * 100, 1) if rev_total else None

    rows, g_total = [], 0
    prev_cat = None
    for grp in order:
        arts = gmap[grp]
        cat  = _get_grp_category(arts)

        if add_cat_headers and cat != prev_cat:
            lbl = 'ПРОИЗВОДСТВЕННЫЕ РАСХОДЫ' if cat == 'Производственные' else 'АДМИНИСТРАТИВНЫЕ РАСХОДЫ'
            rows.append({'type': 'cat_head', 'label': lbl, 'category': cat})
            prev_cat = cat

        g_vals = [float(pv.loc[pv.index.isin(arts), iv].sum()) for iv in intervals]
        g_sum  = sum(g_vals)
        if g_sum == 0:
            continue
        g_total += g_sum
        rows.append({
            'type': 'group', 'label': grp, 'single': len(arts) == 1,
            'category': cat,
            'iv_values': [round(v) for v in g_vals],
            'total': round(g_sum),
            'avg':   round(g_sum / max(len(intervals), 1)),
            'pct_rev': sp(g_sum),
        })
        if len(arts) > 1:
            for art in sorted(arts,
                    key=lambda a: abs(float(pv.loc[a].sum())) if a in pv.index else 0,
                    reverse=True):
                a_vals = [float(pv.loc[art, iv]) if art in pv.index else 0 for iv in intervals]
                a_sum  = sum(a_vals)
                rows.append({
                    'type': 'detail', 'label': art, 'single': False,
                    'category': cat,
                    'iv_values': [round(v) for v in a_vals],
                    'total':   round(a_sum),
                    'avg':     round(a_sum / max(len(intervals), 1)),
                    'pct_rev': sp(a_sum),
                })
    return rows, g_total


def build_period_analysis(date_from, date_to, interval):
    tx = get_tx()
    from_dt = pd.Timestamp(date_from)
    to_dt   = pd.Timestamp(date_to)

    ivs, lbls, iv_fn = _make_interval_col(interval, from_dt, to_dt)

    # Revenue per interval from DDS sheet
    if interval == 'year':
        rev_values = [round(get_revenue(int(iv), list(range(1, 13)))) for iv in ivs]
    elif interval == 'month':
        rev_values = []
        for iv in ivs:
            y, m = int(iv.split('-')[0]), int(iv.split('-')[1])
            rev_values.append(round(get_revenue(y, [m])))
    else:
        rev_values = [0] * len(ivs)   # no daily granularity in DDS

    rev_total = sum(rev_values)

    pv = _build_interval_pivot(tx, date_from, date_to, ivs, iv_fn, credit=False)
    pv = pv[(pv > 0).any(axis=1)]
    expense_rows, exp_total = _build_interval_rows(pv, ivs, rev_total)

    if expense_rows:
        grand_iv = [round(float(pv[iv].sum())) if iv in pv.columns else 0 for iv in ivs]
        expense_rows.append({
            'type': 'grand', 'label': 'ИТОГО РАСХОДЫ', 'single': False,
            'iv_values': grand_iv,
            'total':   round(exp_total),
            'avg':     round(exp_total / max(len(ivs), 1)),
            'pct_rev': round(exp_total / rev_total * 100, 1) if rev_total else None,
        })

    pv_cr = _build_interval_pivot(tx, date_from, date_to, ivs, iv_fn, credit=True)
    credit_rows, cr_total = _build_interval_rows(pv_cr, ivs, rev_total, add_cat_headers=False)
    if credit_rows:
        cr_iv = [round(float(pv_cr[iv].sum())) if iv in pv_cr.columns else 0 for iv in ivs]
        credit_rows.append({
            'type': 'grand', 'label': 'ИТОГО КРЕДИТЫ И ЗАЙМЫ', 'single': False,
            'iv_values': cr_iv,
            'total': round(cr_total),
            'avg':   round(cr_total / max(len(ivs), 1)),
            'pct_rev': None,
        })

    # Revenue row (for display)
    rev_row = {
        'type': 'revenue', 'label': 'ВЫРУЧКА', 'single': False,
        'iv_values': rev_values,
        'total': rev_total,
        'avg':   round(rev_total / max(len(ivs), 1)),
        'pct_rev': None,
    }

    # Разбивка выручки по статьям из транзакций
    rev_breakdown = get_interval_revenue_breakdown(tx, date_from, date_to, ivs, iv_fn)

    return {
        'date_from':       date_from,
        'date_to':         date_to,
        'interval':        interval,
        'intervals':       ivs,
        'interval_labels': lbls,
        'rev_row':         rev_row,
        'rev_total':       rev_total,
        'exp_total':       round(exp_total),
        'rows':            expense_rows,
        'credit_rows':     credit_rows,
        'rev_breakdown':   rev_breakdown,
    }


@app.route('/api/period-analysis', methods=['POST'])
def api_period_analysis():
    data = request.json
    try:
        report = build_period_analysis(
            data['date_from'],
            data['date_to'],
            data.get('interval', 'month'),
        )
        return jsonify(report)
    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/download-period', methods=['POST'])
def api_download_period():
    data = request.json
    try:
        report = build_period_analysis(
            data['date_from'],
            data['date_to'],
            data.get('interval', 'month'),
        )
        buf   = generate_xlsx_period(report)
        fname = f"period_{data['date_from']}_{data['date_to']}.xlsx"
        return send_file(buf, download_name=fname, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/gsheets-status', methods=['GET'])
def api_gsheets_status():
    return jsonify({
        'available':         _GSPREAD_AVAILABLE,
        'credentials_found': CREDENTIALS_FILE.exists(),
        'enabled':           USE_GSHEETS,
        'gsheet_id':         GSHEET_ID,
    })


@app.route('/api/gsheets-sync', methods=['POST'])
def api_gsheets_sync():
    """Принудительная перезагрузка данных из Google Sheets."""
    global USE_GSHEETS
    if not _GSPREAD_AVAILABLE:
        return jsonify({
            'ok': False,
            'error': 'Пакеты gspread/google-auth не установлены. '
                     'Выполните: pip install gspread google-auth',
        }), 400
    if not CREDENTIALS_FILE.exists():
        return jsonify({
            'ok': False,
            'error': f'Файл credentials.json не найден в папке приложения.',
        }), 400
    try:
        USE_GSHEETS = True
        reset_cache()
        get_tx()    # проверочная загрузка
        get_dds()
        return jsonify({
            'ok':       True,
            'message':  'Данные успешно обновлены из Google Sheets.',
            'file_info': get_current_file_info(),
        })
    except Exception as e:
        USE_GSHEETS = False
        return jsonify({'ok': False, 'error': str(e)}), 400


# ── Auto-enable Google Sheets if credentials present ──────────────────────────
def _try_enable_gsheets():
    import os
    global USE_GSHEETS
    has_creds = CREDENTIALS_FILE.exists() or bool(os.environ.get('GOOGLE_CREDENTIALS_JSON'))
    if _GSPREAD_AVAILABLE and has_creds:
        USE_GSHEETS = True
        src = 'credentials.json' if CREDENTIALS_FILE.exists() else 'переменная окружения GOOGLE_CREDENTIALS_JSON'
        print(f'✓ Google Sheets включён ({src}). ID: {GSHEET_ID}')
    else:
        if not _GSPREAD_AVAILABLE:
            print('  gspread не установлен — используется режим загрузки xlsx')
        else:
            print('  credentials.json не найден — используется режим загрузки xlsx')

_try_enable_gsheets()


if __name__ == '__main__':
    if USE_GSHEETS and CREDENTIALS_FILE.exists():
        try:
            get_tx()
            get_dds()
        except Exception as e:
            print(f'  Предзагрузка из Google Sheets не удалась: {e}')
    elif XLSX_IN is not None and XLSX_IN.exists():
        get_tx()
        get_dds()
    app.run(debug=False, port=5000)
